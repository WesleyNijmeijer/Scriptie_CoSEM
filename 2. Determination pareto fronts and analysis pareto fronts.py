# -*- coding: utf-8 -*-
"""
Created on Thu Dec 22 19:02:54 2022

@author: Wesley.Nijmeijer
"""

import csv
import itertools
import matplotlib.pyplot as plt
import networkx as nx
import openpyxl
import time


"""
=============================================================================
Input variables
=============================================================================
"""

#defining the file containing the trajects on which a hydrogen pipeline can be
#constructed
file = r"Example_modelinput.xlsx"

#Defining the year fow which the calculations are done. Either 2030 or 2050 can be chosen. Make sure the input file correlates with the year chosen here.
modelyear = "2030"

#demand scenarios for the 2 years for which the calculations can be done, as the names of the scenarios differ for the 2 options 
if modelyear == "2050":
    demand_scenarios = ["Lowflow", "HighdrogenNL", "Hyconnect", "Rotterdamrules" ]  #Names of the different demand scenarios for 2050
elif modelyear == "2030":
     demand_scenarios = ["Conservative","Middle","Progressive", "Progressive+"]  #Names of the different demand scenarios for 2030

#costs associated with constructing or conversion of a pipeline
c1 = 3200000 #Cost of the construction of 1 km of new pipeline with capacity 200
c2 = 840000  #Cost of the transformation of 1 km of natural gas to hydrogen pipeline with capacity 200
cap = 200    #the capacity that correlates with the cost of construction or conversion of 1km of pipeline mentioned above
Rc = 0.6 #increase in cost if the capacity of a new pipeline doubles

max_row_pipelines = 42     #Last rownumber containing data in the worksheet 'Pipelines' in the Excel input
max_row_nodes= 35          #Last rownumber containing data on the 'regular' nodes (not considering the addiitonal nodes) in the worksheet 'Nodes' in the Excel input
max_row_capacities = 43    #Last rownumber containing data in the worksheet 'Capacities' in the Excel input


#%%
"""
=============================================================================
Timing the calculations
=============================================================================
"""

starttime = time.time()

#%%
"""
=============================================================================
Reading xlsx file containing possible pipeline trajects and plotting graph with all possible trajects
=============================================================================
"""

# Open the Excel file
wb = openpyxl.load_workbook(file, data_only=True)

# Get the worksheet with the edges
ws1 = wb['Pipelines']

# Create a graph
g = nx.MultiDiGraph(keys=True)

# Add edges to the graph and make a dicionary with the lengths of the pipelines and whether or not re-use is possible
length = {}
reuse = [[],[],[]]

for row in ws1.iter_rows(min_row=2,max_row= max_row_pipelines, values_only=True):
    start_node, end_node, keys = row[0], row[1], row[2]
    g.add_edge(start_node, end_node, key = keys)
    length[(start_node, end_node, keys)] = row[6]
    if row[3] == 'Yes' or row[3] == 'Partly':
        reuse[0].append((start_node, end_node, keys))
        reuse[1].append(row[3])
        reuse[2].append(row[4])

# Open the Excel file
wb = openpyxl.load_workbook(file, data_only=True)

# Get the worksheet with the nodes
ws2 = wb['Nodes']
        
#adding coordinates displaying the graph
pos = {}

for row in ws2.iter_rows(min_row=2, values_only=True):
    pos[row[0]] = (row[2], row[3])

nx.set_node_attributes(g, pos, 'pos')

nx.draw(g, pos=nx.get_node_attributes(g, 'pos'), with_labels=True, connectionstyle='arc3, rad=0.1')
plt.show()

E = list(g.edges())

Ek = []
for u, v, key in g.edges(keys=True):
    Ek.append((u,v,key))
        
n = g.number_of_edges()
N = g.number_of_nodes()
F = sorted(list(g.nodes())) #All nodes in an ordered list

print('Er zijn in totaliteit', n ,'mogelijke trajecten. Deze trajecten verbinden', N, 'knopen.')

#Determining trajects which need to be included as they are essential for the network to be connected
EP = []     #list of the indexes of the essential pipelines in list E

for i in F:
    if g.degree(i) == 1: 
        EP += (E.index((u, v)) for u,v in g.in_edges(i))
        EP +=  (E.index((u, v)) for u,v in g.out_edges(i))
        
if EP == []:
    print('er zijn geen essentiële pijpleidingen in de volledige graaf')
else:
    print('de essentiële pijpleidingen in de volledige graaf zijn:', [E[i] for i in EP])  

#%%
"""
=============================================================================
Reading which nodes are sinks and which are sources in the various scenarios
=============================================================================
"""

# Open the Excel file
wb = openpyxl.load_workbook(file, data_only=True)

# Get the worksheet with the nodes
ws2 = wb['Nodes']

def determing_sinkorsource(rownumber):
    demand = {}
    sinks = 0
    sources = 0
    for row in ws2.iter_rows(min_row=2, max_row= max_row_nodes, values_only=True):
        if row[0] in F:
            if row[rownumber] == None:
                demand[row[0]] = 0
            else:
                demand[row[0]] = round(row[rownumber],0)
                if row[rownumber] > 0:
                    sinks += 1
                elif row[rownumber] < 0:
                    sources += 1
    return demand, sinks, sources
        

for i in demand_scenarios:
    if i == demand_scenarios[0]:
        demand_scenario1, sinks_scenario1, sources_scenario1 = determing_sinkorsource(4)
    elif i == demand_scenarios[1]:
        demand_scenario2, sinks_scenario2, sources_scenario2 = determing_sinkorsource(6)
    elif i == demand_scenarios[2]:
        demand_scenario3, sinks_scenario3, sources_scenario3 = determing_sinkorsource(8)
    elif i == demand_scenarios[3]:
        demand_scenario4, sinks_scenario4, sources_scenario4 = determing_sinkorsource(10)
        
for scenario in demand_scenarios:
    if scenario == demand_scenarios[0]:
        print('er zijn',sources_scenario1,'sources en', sinks_scenario1, 'sinks in the variant',demand_scenarios[0])
    elif scenario == demand_scenarios[1]:
        print('er zijn',sources_scenario2,'sources en', sinks_scenario2, 'sinks in the variant', demand_scenarios[1])
    elif scenario == demand_scenarios[2]:
        print('er zijn',sources_scenario3,'sources en', sinks_scenario3, 'sinks in the variant', demand_scenarios[2])
    elif scenario == demand_scenarios[3]:
        print('er zijn',sources_scenario4,'sources en', sinks_scenario4, 'sinks in the variant', demand_scenarios[3])

#%%
"""
=============================================================================
Determining variants with whether or not a pipeline exists
=============================================================================
"""

trajects = []   #list in the order of the edges to dtermine whether or not a pipeline is essential

for i in E:
    if E.index(i) in EP:
        trajects.append([1])
    else:
        trajects.append([0,1])
        
#%%
"""
=============================================================================
Reading the capacities, which are based on the outcomes of the 
maxflowmincost algorthm
=============================================================================
"""

# Open the Excel file
wb = openpyxl.load_workbook(file, data_only=True)

# Get the worksheet with the nodes
ws3 = wb['Capacities']

new_capacities_scenario1 = {}
new_capacities_scenario2 = {}
new_capacities_scenario3 = {}
new_capacities_scenario4 = {}

for scenario in demand_scenarios:
    for row in ws3.iter_rows(min_row=3,max_row= max_row_capacities, values_only=True):
        if scenario == demand_scenarios[0]:
            if row[4] > 0:
                new_capacities_scenario1[(row[0]),row[1],row[2]] = row[3],row[4]
            else:
                new_capacities_scenario1[(row[0]),row[1],row[2]] = row[3]
        elif scenario == demand_scenarios[1]:
            if row[6] > 0:
                new_capacities_scenario2[(row[0]),row[1],row[2]] = row[5],row[6]
            else:
                new_capacities_scenario2[(row[0]),row[1],row[2]] = row[6]
        elif scenario == demand_scenarios[2]:
            if row[8] > 0:
                new_capacities_scenario3[(row[0]),row[1],row[2]] = row[7],row[8]
            else:
                new_capacities_scenario3[(row[0]),row[1],row[2]] = row[8]
        elif scenario == demand_scenarios[3]:
            if row[10] > 0:
                new_capacities_scenario4[(row[0]),row[1],row[2]] = row[9],row[10]
            else:
                new_capacities_scenario4[(row[0]),row[1],row[2]] = row[9]

#%%
"""
=============================================================================
Determining variants and calculating costst of and maximal flows through the network
=============================================================================
"""

filename = 'results'
filenumber = '0'

counter = 0

filesave = open((filename+filenumber+'.csv'), 'w', newline='')
writer = csv.writer(filesave, delimiter=',')
writer.writerow(["scenario", "v","c", "flow_value", "costs", "flow_dict"])

x_scenario1 = []
x_scenario2 = []
x_scenario3 = []
x_scenario4 = []

y_scenario1 = []
y_scenario2 = []
y_scenario3 = []
y_scenario4 = []

def set_capacities(GE,new_capacities):
    capacities = []
    for edge in GE:
        if type(new_capacities[edge]) == int:
            capacities.append([0,3])
        elif len(new_capacities[edge]) == 2:
            capacities.append([0,1,3])
    return capacities
      
def calculate_costs(c,cap_e):
    # Calculating costs of the network
    costs = 0
    for i in range(n1):
        if (GE[i][0],GE[i][1],GE[i][2]) in reuse[0]:
            index = reuse[0].index((GE[i][0],GE[i][1],GE[i][2]))
            if cap_e[(GE[i][0],GE[i][1],GE[i][2])] >= reuse[2][index]:
                costs += length[(GE[i][0],GE[i][1],GE[i][2])]*c2*((reuse[2][index])/cap)**Rc
                costs += length[(GE[i][0],GE[i][1],GE[i][2])]*c1*((cap_e[(GE[i][0],GE[i][1],GE[i][2])]-reuse[2][index])/cap)**Rc
            else:
                costs += length[(GE[i][0],GE[i][1],GE[i][2])]*c1*(cap_e[(GE[i][0],GE[i][1],GE[i][2])]/cap)**Rc
        elif (GE[i][1],GE[i][0],GE[i][2]) in reuse[0]:
            index = reuse[0].index(((GE[i][1],GE[i][0],GE[i][2])))
            if cap_e[(GE[i][1],GE[i][0],GE[i][2])] >= reuse[2][index]:
                costs += length[(GE[i][0],GE[i][1],GE[i][2])]*c2*((reuse[0][index])/cap)**Rc
                costs += length[(GE[i][0],GE[i][1],GE[i][2])]*c1*((cap_e[(GE[i][0],GE[i][1],GE[i][2])]-reuse[0][index])/cap)**Rc
            else:
                costs += length[(GE[i][0],GE[i][1],GE[i][2])]*c1*(cap_e[(GE[i][0],GE[i][1],GE[i][2])]/cap)**Rc
        else:
            costs += length[(GE[i][0],GE[i][1],GE[i][2])]*c1*(cap_e[(GE[i][0],GE[i][1],GE[i][2])]/cap)**Rc
    costs = round((costs/1000000),1)
    return costs

def calculate_maxflow(G,scenario, demand):
    #as maximal flow cannot be calculated capacities are added for routes where 2 edges exist
    H = nx.Graph()
    for o,p,key,data in G.edges(data=True, keys=True):
        if (o,p) not in H.edges():
            H.add_edge(o, p, capacity=data['capacity'])
        else:
            H[o][p]['capacity'] += data['capacity']
    # Make directed graph
    I = H.to_directed()
    # Apply maximum flow to find the maximum flow through the network
    I.add_nodes_from(['supersource', 'supersink'])
    for i in F:
        if int(demand[i]) < 0:
            I.add_edge('supersource',i, capacity = -int(demand[i]))
        elif int(demand[i]) > 0:
            I.add_edge(i, 'supersink', capacity = int(demand[i]))
    flow_value, flow_dict = nx.maximum_flow(I, 'supersource', 'supersink')
    return flow_value, flow_dict 

for v in itertools.product(*trajects):
    n1 = v.count(1) #number of edges in graph
    if n1 >= N-1:
        G = nx.MultiGraph([Ek[i] for i in range(len(v)) if v[i]== 1]) #Define graph
        G.add_nodes_from(F) #Add all nodes
        l_N = list(G.nodes())
        if nx.is_connected(G):
            GE = [Ek[i] for i in range(n) if v[i] == 1]
            # Length of edges 
            nx.set_edge_attributes(G,length, 'length')
            
            for scenario in demand_scenarios:
                
                if scenario == demand_scenarios[0]:
                    capacities_scenario1 = set_capacities(GE,new_capacities_scenario1)
                    for c in itertools.product(*capacities_scenario1):
                        counter += 1
                        if counter % 1000000 == 0:
                            filesave.close()
                            filenumber = str(int(counter/1000000))
                            filesave = open((filename+filenumber+'.csv'), 'w', newline='')
                            writer = csv.writer(filesave, delimiter=',')
                        cap_e_scenario1 = {}
                        for i in range(n1):
                            if c[i] == 0:
                                if type(new_capacities_scenario1[GE[i]]) == int:
                                    cap_e_scenario1[GE[i]] = new_capacities_scenario1[GE[i]]
                                else:
                                   cap_e_scenario1[GE[i]] = new_capacities_scenario1[GE[i]][0]
                            else:
                                cap_e_scenario1[GE[i]] = new_capacities_scenario1[GE[i]][1]
                        G1 = G.copy()
                        nx.set_edge_attributes(G1, cap_e_scenario1, 'capacity')
                        costs = calculate_costs(c,cap_e_scenario1)
                        x_scenario1.append(costs)
                        flow_value, flow_dict = calculate_maxflow(G1,scenario, demand_scenario1)
                        y_scenario1.append(round(flow_value))
                        writer.writerow([scenario,v,c,round(flow_value), costs,flow_dict])
                        
                elif scenario == demand_scenarios[1]:
                    capacities_scenario2 = set_capacities(GE,new_capacities_scenario2)
                    for c in itertools.product(*capacities_scenario2):
                        counter += 1
                        if counter % 1000000 == 0:
                            filesave.close()
                            filenumber = str(int(counter/1000000))
                            filesave = open((filename+filenumber+'.csv'), 'w', newline='')
                            writer = csv.writer(filesave, delimiter=',')
                        cap_e_scenario2 = {}
                        for i in range(n1):
                            if c[i] == 0:
                                if type(new_capacities_scenario2[GE[i]]) == int:
                                    cap_e_scenario2[GE[i]] = new_capacities_scenario2[GE[i]]
                                else:
                                      cap_e_scenario2[GE[i]] = new_capacities_scenario2[GE[i]][0]
                            else:
                                cap_e_scenario2[GE[i]] = new_capacities_scenario2[GE[i]][1]
                        G1 = G.copy()
                        nx.set_edge_attributes(G1, cap_e_scenario2, 'capacity')
                        costs = calculate_costs(c,cap_e_scenario2)
                        x_scenario2.append(costs)
                        flow_value, flow_dict = calculate_maxflow(G1,scenario, demand_scenario2)
                        y_scenario2.append(round(flow_value))
                        writer.writerow([scenario,v,c,round(flow_value), costs,flow_dict])
                            
                elif scenario == demand_scenarios[2]:
                    capacities_scenario3 = set_capacities(GE,new_capacities_scenario3)
                    for c in itertools.product(*capacities_scenario3):
                        counter += 1
                        if counter % 1000000 == 0:
                            filesave.close()
                            filenumber = str(int(counter/1000000))
                            filesave = open((filename+filenumber+'.csv'), 'w', newline='')
                            writer = csv.writer(filesave, delimiter=',')
                        cap_e_scenario3 = {}
                        for i in range(n1):
                            if c[i] == 0:
                                if type(new_capacities_scenario3[GE[i]]) == int:
                                    cap_e_scenario3[GE[i]] = new_capacities_scenario3[GE[i]]
                                else:
                                    cap_e_scenario3[GE[i]] = new_capacities_scenario3[GE[i]][0]
                            else:
                                cap_e_scenario3[GE[i]] = new_capacities_scenario3[GE[i]][1]
                        G1 = G.copy()
                        nx.set_edge_attributes(G1, cap_e_scenario3, 'capacity')
                        costs = calculate_costs(c,cap_e_scenario3)
                        x_scenario3.append(costs)
                        flow_value,flow_dict = calculate_maxflow(G1,scenario, demand_scenario3)
                        y_scenario3.append(round(flow_value))
                        writer.writerow([scenario,v,c,round(flow_value), costs,flow_dict])
                
                elif scenario == demand_scenarios[3]:
                    capacities_scenario4 = set_capacities(GE,new_capacities_scenario4)
                    for c in itertools.product(*capacities_scenario4):
                        counter += 1
                        if counter % 1000000 == 0:
                            filesave.close()
                            filenumber = str(int(counter/1000000))
                            filesave = open((filename+filenumber+'.csv'), 'w', newline='')
                            writer = csv.writer(filesave, delimiter=',')
                        cap_e_scenario4 = {}
                        for i in range(n1):
                            if c[i] == 0:
                                if type(new_capacities_scenario4[GE[i]]) == int:
                                    cap_e_scenario4[GE[i]] = new_capacities_scenario4[GE[i]]
                                else:
                                   cap_e_scenario4[GE[i]] = new_capacities_scenario4[GE[i]][0]
                            else:
                                cap_e_scenario4[GE[i]] = new_capacities_scenario4[GE[i]][1]
                        G1 = G.copy()
                        nx.set_edge_attributes(G1, cap_e_scenario4, 'capacity')
                        costs = calculate_costs(c,cap_e_scenario4)
                        x_scenario4.append(costs)           
                        flow_value, flow_dict = calculate_maxflow(G1,scenario, demand_scenario4)
                        y_scenario4.append(round(flow_value))
                        writer.writerow([scenario,v,c,round(flow_value), costs, flow_dict])
        
filesave.close()

#%%
"""
=============================================================================
Calculating the pareto front
=============================================================================
"""

def calculating_pareto(scenario,x,y):
    sorted_list = sorted([[x[i], y[i]] for i in range(len(x))])
    start = [sorted_list[0]]
    
    for pair in sorted_list:
        if pair[0] == start[0][0]:
            if pair[1] > start[0][1]:
                    start.clear()
                    start.append(pair) 
    
    
    pareto_front = []
    for pair in sorted_list:
        if pair == start[0]:
            pareto_front.append(pair)
        elif pair[1] > start[0][1]:
            if pair[1] >= pareto_front[-1][1]:
                if pair[0] == pareto_front[-1][0] and pair != pareto_front[-1]:
                    pareto_front.remove(pareto_front[-1])
                    pareto_front.append(pair)
                elif pair == pareto_front[-1]:
                    pareto_front.append(pair)
                elif pair[0] >= pareto_front[-1][0] and pair[1] <= pareto_front[-1][1]:
                    pass
                else:
                    pareto_front.append(pair)                
    return pareto_front
    
def plot_pareto(x,y,scenario, pareto_front):
    plt.plot(y, x, '.b', markersize=16, label='Non Pareto-optimal')
    pf_x = [pair[0] for pair in pareto_front]
    pf_y = [pair[1] for pair in pareto_front]
    plt.plot(pf_y, pf_x, '.r', markersize=16, label=('Pareto optimal'))
    plt.title(scenario, fontsize=24)
    plt.ylabel("Costs (M€)", fontsize=16)
    plt.xlabel("Maximum flow (PJ)", fontsize =16)
    plt.legend(loc='upper left', numpoints=1)
    plt.show()

for scenario in demand_scenarios:
    if scenario == demand_scenarios[0]:
        pareto_scenario1 = calculating_pareto(scenario,x_scenario1,y_scenario1)
        plot_pareto(x_scenario1,y_scenario1,scenario,pareto_scenario1)
    elif scenario == demand_scenarios[1]:
        pareto_scenario2 = calculating_pareto(scenario,x_scenario2,y_scenario2)
        plot_pareto(x_scenario2,y_scenario2,scenario,pareto_scenario2)
    elif scenario == demand_scenarios[2]:
        pareto_scenario3 = calculating_pareto(scenario,x_scenario3,y_scenario3)
        plot_pareto(x_scenario3,y_scenario3,scenario,pareto_scenario3)
    elif scenario == demand_scenarios[3]:
        pareto_scenario4 = calculating_pareto(scenario,x_scenario4,y_scenario4)
        plot_pareto(x_scenario4,y_scenario4,scenario,pareto_scenario4)
        
#%%
"""
=============================================================================
Plotting the pareto fronts of the different scenarios in one plot, so that
the different variants can be judged on the different scenarios
=============================================================================
"""

for scenario in demand_scenarios:
    if scenario == demand_scenarios[0]:
        pf_x_scenario1 = [pair[0] for pair in pareto_scenario1]
        pf_y_scenario1 = [pair[1] for pair in pareto_scenario1]
        plt.plot(pf_y_scenario1, pf_x_scenario1, '.b', markersize=16, label=(demand_scenarios[0]))
    elif scenario == demand_scenarios[1]:
        pf_x_scenario2 = [pair[0] for pair in pareto_scenario2]
        pf_y_scenario2 = [pair[1] for pair in pareto_scenario2]
        plt.plot(pf_y_scenario2, pf_x_scenario2, '.g', markersize=16, label=(demand_scenarios[1]))
    elif scenario == demand_scenarios[2]:
        pf_x_scenario3 = [pair[0] for pair in pareto_scenario3]
        pf_y_scenario3 = [pair[1] for pair in pareto_scenario3]
        plt.plot(pf_y_scenario3, pf_x_scenario3, '.r', markersize=16, label=(demand_scenarios[2]))
    elif scenario == demand_scenarios[3]:
        pf_x_scenario4 = [pair[0] for pair in pareto_scenario4]
        pf_y_scenario4 = [pair[1] for pair in pareto_scenario4]
        plt.plot(pf_y_scenario4, pf_x_scenario4, '.m', markersize=16, label=(demand_scenarios[3]))
plt.title("pareto fronts scenarios", fontsize=24)
plt.ylabel("Costs (M€)", fontsize=16)
plt.xlabel("Maximum flow (PJ)", fontsize =16)
plt.legend(loc='lower right', numpoints=1)
plt.show()

#%%
"""
=============================================================================
Creating an excel file with whether or not a certain variant is part of 
the pareto front
=============================================================================
"""

def change_column(flow_value, costs, pareto_front):
    if [costs, flow_value] in pareto_front:
        return 'True'
    else:
        return 'False'
    
def creating_csv_pareto(filenumber,pareto_filename, scenario, pareto_front):
    with open((filename+filenumber+'.csv'), 'r') as f_input:
        reader = csv.reader(f_input)
        next(reader, None)
        for row in reader:
            if row[0] == scenario:
                newrow = [row[0], row[1], row[2], row[3],row[4],row[5]]
                flow_value = float(row[3])
                costs = float(row[4])
                new_value = change_column(flow_value, costs, pareto_front)
                newrow.append(new_value)
                if new_value == 'True':
                    with open(pareto_filename, 'a', newline='') as f_output:
                        writer = csv.writer(f_output)
                        writer.writerow(newrow)

def headers_csv_pareto(scenario, pareto_filename):
    with open((filename+str(0)+'.csv'), 'r') as f_input:
        reader = csv.reader(f_input)
        headers = next(reader)
        headers.append('Pareto front')
        headers.append('location')
        headers.append(Ek)

    with open(pareto_filename, 'a', newline='') as f_output:
        writer = csv.writer(f_output)
        writer.writerow(headers)

for scenario in demand_scenarios:
    if scenario == demand_scenarios[0]:
        pareto_filename = 'results_pareto' + scenario + '.csv'
        headers_csv_pareto(scenario, pareto_filename)
        if filenumber == '0':
            creating_csv_pareto(filenumber,pareto_filename, scenario, pareto_scenario1)
        else:
            for i in range(int(float(filenumber))):
                creating_csv_pareto(str(i), pareto_filename, scenario, pareto_scenario1)
    elif scenario == demand_scenarios[1]:
        pareto_filename = 'results_pareto' + scenario + '.csv'
        headers_csv_pareto(scenario, pareto_filename)
        if filenumber == '0':
            creating_csv_pareto(filenumber,pareto_filename,scenario, pareto_scenario2)
        else:
            for i in range(int(float(filenumber))):
                creating_csv_pareto(str(i),pareto_filename,scenario, pareto_scenario2)      
    elif scenario == demand_scenarios[2]:
        pareto_filename = 'results_pareto' + scenario + '.csv'
        headers_csv_pareto(scenario, pareto_filename)
        if filenumber == '0':
            creating_csv_pareto(filenumber,pareto_filename,scenario, pareto_scenario3)
        else:
            for i in range(int(float(filenumber))):
                creating_csv_pareto(str(i),pareto_filename,scenario, pareto_scenario3)
    elif scenario == demand_scenarios[3]:
        pareto_filename = 'results_pareto' + scenario + '.csv'
        headers_csv_pareto(scenario, pareto_filename)
        if filenumber == '0':
            creating_csv_pareto(filenumber,pareto_filename,scenario, pareto_scenario4)
        else:
            for i in range(int(float(filenumber))):
                creating_csv_pareto(str(i),pareto_filename,scenario, pareto_scenario4)   

#%%
"""
=============================================================================
Determining whether or not a certain pipeline is part of the entire pareto 
front or not
=============================================================================
"""

def cleanup_string(string):
    clean ="".join(c for c in string if c.isalnum())
    return clean

def determine_existence(pareto_filename):
    existence = {}
    with open(pareto_filename, 'r', newline='') as f_input:
        reader = csv.reader(f_input)
        next(reader, None)
        firstrow = next(reader, None)
        vstart = cleanup_string(firstrow[1])
        for row in reader:
            v = cleanup_string(row[1])
            for i in range(n):
                if v[i] == vstart[i]:
                    if Ek[i] in existence:
                        if existence[Ek[i]] == 'Unclear':
                            pass
                        else:
                            if int(v[i]) == 1:
                                existence[Ek[i]] = 'Yes'
                            else:
                                existence[Ek[i]] = 'No'
                    else:
                        if int(v[i]) == 1:
                            existence[Ek[i]] = 'Yes'
                        else:
                            existence[Ek[i]] = 'No'
                else:
                    existence[Ek[i]] = 'Unclear'
    return existence

def plot_network(existence, scenario,demand):
    edges_red = []
    edges_yellow = []
    for a,b,c in list(existence.keys()):
        if existence[(a,b,0)] == 'No':
            edges_red.append((a,b))
        if existence[(a,b,0)] == 'Unclear':
            edges_yellow.append((a,b))
    G = nx.Graph()
    E_reduced = []
    for edge in E:
        if edge not in E_reduced:
            E_reduced.append(edge)
    G.add_edges_from(E_reduced)
    G.add_nodes_from(F)
    node_color = []
    for node in G.nodes():
        if demand[node] == 0:
            node_color.append('blue')
        elif demand[node] > 0:
            node_color.append('orange')
        elif demand[node] < 0:
            node_color.append('green')
    plt.figure(figsize=(15, 15))
    plt.title((modelyear+","+scenario),size=50.0)
    node_patches =[]
    node_patches.append(plt.plot([],[], marker="o", ms=10, ls="", mec=None, color= 'green', label="Source", alpha=0.5)[0])
    node_patches.append(plt.plot([],[], marker="o", ms=10, ls="", mec=None, color='orange', label="Sink", alpha=0.5)[0])
    node_patches.append(plt.plot([],[], marker="o", ms=10, ls="", mec=None, color='blue', label="Neither", alpha=0.5)[0])
    edge_patches = []
    edge_patches.append(plt.plot([],[], lw=2, color='red', label="Pipeline unused", alpha=0.5)[0])
    edge_patches.append(plt.plot([],[], lw=2, color='yellow', label="Pipeline variant dependant", alpha=0.5)[0])    
    edge_patches.append(plt.plot([],[], lw=2, color='black', label="Pipeline used", alpha=0.5)[0])
    plt.legend(handles=node_patches + edge_patches, loc='lower right', fontsize = 15, markerscale= 2)
    nx.draw_networkx(G,pos=pos,with_labels=True,width=5.0,node_size= 600,node_color=node_color,font_size=20,font_color='white')
    nx.draw_networkx_edges(G,pos=pos,edgelist=(edges_red),width=3.0,edge_color='red')
    nx.draw_networkx_edges(G,pos=pos,edgelist=(edges_yellow),width=5.0,edge_color='yellow')
    plt.show()   

for scenario in demand_scenarios:
    if scenario == demand_scenarios[0]:
        existence_scenario1 = determine_existence(('results_pareto' + scenario + '.csv'))
        plot_network(existence_scenario1, scenario,demand_scenario1)
    elif scenario == demand_scenarios[1]:
        existence_scenario2 = determine_existence(('results_pareto' + scenario + '.csv'))
        plot_network(existence_scenario2, scenario,demand_scenario2)
    elif scenario == demand_scenarios[2]:
        existence_scenario3 = determine_existence(('results_pareto' + scenario + '.csv'))
        plot_network(existence_scenario3, scenario,demand_scenario3)
    elif scenario == demand_scenarios[3]:
        existence_scenario4 = determine_existence(('results_pareto' + scenario + '.csv'))
        plot_network(existence_scenario4, scenario,demand_scenario4)

#%%
"""
=============================================================================
Determining the maximal flow of the different pareto fronts on the other 
demand scenarios
=============================================================================
"""
        
filesave = open(('analysispareto'+modelyear+'.csv'), 'w', newline='')
writer = csv.writer(filesave, delimiter=',')
writer.writerow(["original scenario","v","c","costs",("flow_value_" + demand_scenarios[0]),("flow_value_" + demand_scenarios[1]), ("flow_value_" + demand_scenarios[2]), ("flow_value_" + demand_scenarios[3])])

def determine_cap_e(c,GE, new_capacities):
    cap_e = {}
    for i in range(len(c)):
        if c[i] == '0':
            if type(new_capacities[GE[i]]) == int:
                cap_e[GE[i]] = new_capacities[GE[i]]
            else:
                  cap_e[GE[i]] = new_capacities[GE[i]][0]
        else:
            cap_e[GE[i]] = new_capacities[GE[i]][1]
    return cap_e

def calculate_maxflow(G,demand):
    #as maximal flow cannot be calculated capacities are added for routes where 2 edges exist
    H = nx.Graph()
    for o,p,key,data in G.edges(data=True, keys=True):
        if (o,p) not in H.edges():
            H.add_edge(o, p, capacity=data['capacity'])
        else:
            H[o][p]['capacity'] += data['capacity']
    # Make directed graph
    I = H.to_directed()
    # Apply maximum flow to find the maximum flow through the network
    I.add_nodes_from(['supersource', 'supersink'])
    for i in F:
        if int(demand[i]) < 0:
            I.add_edge('supersource',i, capacity = -int(demand[i]))
        elif int(demand[i]) > 0:
            I.add_edge(i, 'supersink', capacity = int(demand[i]))
    flow_value, flow_dict = nx.maximum_flow(I, 'supersource', 'supersink')
    return flow_value, flow_dict 



def determine_performance_on_other_scenarios(filenamepareto):
    with open(filenamepareto, 'r') as f_input:
        reader = csv.reader(f_input)
        next(reader, None)
        for row in reader:
            scenario = row[0]
            v = cleanup_string(row[1])
            c = cleanup_string(row[2])
            costs = float(row[4])
            
            K = nx.MultiGraph([Ek[i] for i in range(len(v)) if v[i]== '1'])
            GE = [Ek[i] for i in range(n) if v[i] == '1']
            
            if scenario == demand_scenarios[0]:
                flow_value_scenario1 = float(row[3])
                
                cap_e = determine_cap_e(c,GE, new_capacities_scenario1)
                nx.set_edge_attributes(K, cap_e, 'capacity')
                
                flow_value_scenario2, flowdict = calculate_maxflow(K, demand_scenario2)
                flow_value_scenario3, flowdict = calculate_maxflow(K, demand_scenario3)
                flow_value_scenario4, flowdict = calculate_maxflow(K, demand_scenario4)
             
                writer.writerow([scenario,v,c, costs,flow_value_scenario1,flow_value_scenario2,flow_value_scenario3,flow_value_scenario4])

            if scenario == demand_scenarios[1]:
                flow_value_scenario2 = float(row[3])
                
                cap_e = determine_cap_e(c,GE, new_capacities_scenario2)
                nx.set_edge_attributes(K, cap_e, 'capacity')
                
                flow_value_scenario1, flowdict = calculate_maxflow(K, demand_scenario1)
                flow_value_scenario3, flowdict = calculate_maxflow(K, demand_scenario3)
                flow_value_scenario4, flowdict = calculate_maxflow(K, demand_scenario4)
             
                writer.writerow([scenario,v,c, costs,flow_value_scenario1,flow_value_scenario2,flow_value_scenario3,flow_value_scenario4])

            if scenario == demand_scenarios[2]:
                flow_value_scenario3 = float(row[3])
                
                cap_e = determine_cap_e(c,GE, new_capacities_scenario3)
                nx.set_edge_attributes(K, cap_e, 'capacity')
                
                flow_value_scenario1, flowdict = calculate_maxflow(K, demand_scenario1)
                flow_value_scenario2, flowdict = calculate_maxflow(K, demand_scenario2)
                flow_value_scenario4, flowdict = calculate_maxflow(K, demand_scenario4)
             
                writer.writerow([scenario,v,c, costs,flow_value_scenario1,flow_value_scenario2,flow_value_scenario3,flow_value_scenario4])
                
            if scenario == demand_scenarios[3]:
                flow_value_scenario4 = float(row[3])
                
                cap_e = determine_cap_e(c,GE, new_capacities_scenario4)
                nx.set_edge_attributes(K, cap_e, 'capacity')
                
                flow_value_scenario1, flowdict = calculate_maxflow(K, demand_scenario1)
                flow_value_scenario2, flowdict = calculate_maxflow(K, demand_scenario2)
                flow_value_scenario3, flowdict = calculate_maxflow(K, demand_scenario3)
             
                writer.writerow([scenario,v,c, costs,flow_value_scenario1,flow_value_scenario2,flow_value_scenario3,flow_value_scenario4])


for scenario in demand_scenarios:
    if scenario == demand_scenarios[0]:
        filenamepareto = 'results_pareto' + scenario + '.csv'
        determine_performance_on_other_scenarios(filenamepareto)
    elif scenario == demand_scenarios[1]:
        filenamepareto = 'results_pareto' + scenario + '.csv'
        determine_performance_on_other_scenarios(filenamepareto)
    elif scenario == demand_scenarios[2]:
        filenamepareto = 'results_pareto' + scenario + '.csv'
        determine_performance_on_other_scenarios(filenamepareto)
    elif scenario == demand_scenarios[3]:
        filenamepareto = 'results_pareto' + scenario + '.csv'
        determine_performance_on_other_scenarios(filenamepareto)
    
filesave.close()
        
                
#%%
"""
=============================================================================
Plotting the results of the determination of the maximal flow of the 
different pareto fronts on the other demand scenarios
=============================================================================
"""

scenario1_scenario1 = []
scenario1_scenario2 = []
scenario1_scenario3 = []
scenario1_scenario4 = []

scenario2_scenario1 = []
scenario2_scenario2 = []
scenario2_scenario3 = []
scenario2_scenario4 = []

scenario3_scenario1 = []
scenario3_scenario2 = []
scenario3_scenario3 = []
scenario3_scenario4 = []

scenario4_scenario1 = []
scenario4_scenario2 = []
scenario4_scenario3 = []
scenario4_scenario4 = []

with open(('analysispareto'+modelyear+'.csv'), 'r') as f_input:
    reader = csv.reader(f_input)
    next(reader, None)
    for row in reader:
        original_scenario = row[0]
        if original_scenario == demand_scenarios[0]:
            scenario1_scenario1.append([float(row[3]),float(row[4])])
            scenario1_scenario2.append([float(row[3]),float(row[5])])
            scenario1_scenario3.append([float(row[3]),float(row[6])])
            scenario1_scenario4.append([float(row[3]),float(row[7])])
        elif original_scenario == demand_scenarios[1]:
            scenario2_scenario1.append([float(row[3]),float(row[4])])
            scenario2_scenario2.append([float(row[3]),float(row[5])])
            scenario2_scenario3.append([float(row[3]),float(row[6])])
            scenario2_scenario4.append([float(row[3]),float(row[7])])
        elif original_scenario == demand_scenarios[2]:
            scenario3_scenario1.append([float(row[3]),float(row[4])])
            scenario3_scenario2.append([float(row[3]),float(row[5])])
            scenario3_scenario3.append([float(row[3]),float(row[6])])
            scenario3_scenario4.append([float(row[3]),float(row[7])])
        elif original_scenario == demand_scenarios[3]:
            scenario4_scenario1.append([float(row[3]),float(row[4])])
            scenario4_scenario2.append([float(row[3]),float(row[5])])
            scenario4_scenario3.append([float(row[3]),float(row[6])])
            scenario4_scenario4.append([float(row[3]),float(row[7])])

def plot_analysis(scenario1,scenario2,scenario3,scenario4,title,legendtitle):
    pf_x = [pair[0] for pair in scenario1]
    pf_y = [pair[1] for pair in scenario1]
    plt.plot(pf_y, pf_x, '.b', markersize=16, label=(demand_scenarios[0]))
    pf_x = [pair[0] for pair in scenario2]
    pf_y = [pair[1] for pair in scenario2]
    plt.plot(pf_y, pf_x, '.g', markersize=16, label=(demand_scenarios[1]))
    pf_x = [pair[0] for pair in scenario3]
    pf_y = [pair[1] for pair in scenario3]
    plt.plot(pf_y, pf_x, '.r', markersize=16, label=(demand_scenarios[2]))
    pf_x = [pair[0] for pair in scenario4]
    pf_y = [pair[1] for pair in scenario4]
    plt.plot(pf_y, pf_x, '.m', markersize=16, label=(demand_scenarios[3]))
    
    plt.title(title, fontsize=24)
    plt.ylabel("Costs (M€)", fontsize=16)
    plt.xlabel("Maximum flow (PJ)", fontsize =16)
    plt.legend(loc='best', numpoints=1,title=legendtitle)
    plt.show()

for scenario in demand_scenarios:
    if scenario == demand_scenarios[0]:
        plot_analysis(scenario1_scenario1,scenario1_scenario2,scenario1_scenario3,scenario1_scenario4,('Pareto front '+scenario+' on scenarios'),'Demand scenario')
        plot_analysis(scenario1_scenario1,scenario2_scenario1,scenario3_scenario1,scenario4_scenario1,('Pareto fronts on scenario '+scenario),'Original variant scenario')
    elif scenario == demand_scenarios[1]:
        plot_analysis(scenario2_scenario1,scenario2_scenario2,scenario2_scenario3,scenario2_scenario4,('Pareto front '+scenario+' on demand scenarios'),'Demand scenario')
        plot_analysis(scenario1_scenario2,scenario2_scenario2,scenario3_scenario2,scenario4_scenario2,('Pareto fronts on scenario '+scenario),'Original variant scenario')
    elif scenario == demand_scenarios[2]:
        plot_analysis(scenario3_scenario1,scenario3_scenario2,scenario3_scenario3,scenario3_scenario4,('Pareto front '+scenario+' on scenarios'),'Demand scenario')
        plot_analysis(scenario1_scenario3,scenario2_scenario3,scenario3_scenario3,scenario4_scenario3,('Pareto fronts on scenario '+scenario),'Original variant scenario')
    elif scenario == demand_scenarios[3]:
        plot_analysis(scenario4_scenario1,scenario4_scenario2,scenario4_scenario3,scenario4_scenario4,('Pareto front '+scenario+' on scenarios'),'Demand scenario')
        plot_analysis(scenario1_scenario4,scenario2_scenario4,scenario3_scenario4,scenario4_scenario4,('Pareto fronts on scenario '+scenario),'Original variant scenario')
        
#%%
"""
=============================================================================
Timing the calculations and make a sound
=============================================================================
"""

endtime = time.time()
print('de berekeningen duurden', round((endtime-starttime)), 'seconden.')

#make a sound when done
print('\a')