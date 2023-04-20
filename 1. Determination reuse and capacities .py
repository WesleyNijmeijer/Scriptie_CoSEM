# -*- coding: utf-8 -*-
"""
Created on Thu Dec 22 19:02:54 2022

@author: Wesley.Nijmeijer
"""

import ast
import csv
import itertools
import matplotlib.pyplot as plt
import math
import networkx as nx
import openpyxl
import time

"""
=============================================================================
Input variables
=============================================================================
"""

#defining the file containing the trajects on which a hydrogen pipeline can be
#constructed
file = r"Example_modelinput.xlsx"

#Defining the year fow which the calculations are done. Either 2030 or 2050 can be chosen. Make sure the input file correlates with the year chosen here.
modelyear = "2030"

#demand scenarios for the 2 years for which the calculations can be done, as the names of the scenarios differ for the 2 options 
if modelyear == "2050":
    demand_scenarios = ["Lowflow", "HighdrogenNL", "Hyconnect", "Rotterdamrules" ]  #Names of the different demand scenarios for 2050
elif modelyear == "2030":
    demand_scenarios = ["Conservative","Middle","Progressive", "Progressive+"]  #Names of the different demand scenarios for 2030

#costs associated with constructing or conversion of a pipeline
c1 = 3200000 #Cost of the construction of 1 km of new pipeline with capacity 200
c2 = 840000  #Cost of the transformation of 1 km of natural gas to hydrogen pipeline with capacity 200
cap = 200    #the capacity that correlates with the cost of construction or conversion of 1km of pipeline mentioned above
Rc = 0.6 #increase in cost if the capacity of a new pipeline doubles


switch_determining_newpipelines_necessary = 'No'    #Switch to determine wheter this model is used to determine the reuse capapcities of the pipelines are sufficient ("Yes") or the model is used to determine the capacities of the different pipelines to be used in further steps of the model

if switch_determining_newpipelines_necessary == 'Yes':
    switch_determining_capacitiesmodel2 = 'No'
else: switch_determining_capacitiesmodel2 = 'Yes'

switch_test = 'No'  # 'Yes' or 'No'. If Yes, the model is only run for the first demand scenario. This creates faster results, so that model improvements can be analysed.

max_row_pipelines = 42              #Last rownumber containing data in the worksheet 'Pipelines' in the Excel input
max_row_nodes= 35                   #Last rownumber containing data on the 'regular' nodes (not considering the addiitonal nodes) in the worksheet 'Nodes' in the Excel input
max_row_capacities = 43             #Last rownumber containing data in the worksheet 'Capacities' in the Excel input
max_row_additionalpipelines = 5     #Last rownumber containing data in the worksheet 'Additional_pipelines' in the Excel input


#%%
"""
=============================================================================
Timing the calculations
=============================================================================
"""

starttime = time.time()     #timer used for timing the calculations. Using the timer different calculations can be compared on their efficiency

#%%
"""
=============================================================================
Reading xlsx file containing possible pipeline trajects and plotting graph 
with all possible trajects
=============================================================================
"""

# Open the Excel file
wb = openpyxl.load_workbook(file, data_only=True)

# Get the worksheet with the edges
ws1 = wb['Pipelines']

# Create an empty graph which contains directed edges and can contain multiple edges on the same route
g = nx.MultiDiGraph(keys=True)

capacities = {}
# Add edges to the graph and make a dictionary with the lengths of the pipelines and whether or not re-use is possible
length = {}
reuse = [[],[],[]]

for row in ws1.iter_rows(min_row=2, max_row= max_row_pipelines, values_only=True):
    start_node, end_node, keys = row[0], row[1], row[2]
    g.add_edge(start_node, end_node, key = keys)
    capacities[(start_node, end_node, keys)] = math.inf
    if row[3] == 'Yes' or row[3] == 'Partly':
        reuse[0].append((start_node, end_node, keys))
        reuse[1].append(row[3])
        reuse[2].append(row[4])
        if switch_determining_capacitiesmodel2 == "Yes":
            capacities[(start_node, end_node, keys)] = (10*row[4])
    if switch_determining_newpipelines_necessary == "Yes" or switch_determining_capacitiesmodel2 == 'Yes':
        if (start_node, end_node, keys) in reuse[0]:
            length[(start_node, end_node, keys)] = int(0.26*row[6])      #Multiplying the length of reused natural gas pipelines by 0.26 to incorporate the fact that reusing natural gas pipelines is cheaper than construction of new pipelines
        else:
            length[(start_node, end_node, keys)] = row[6]
    else:
        length[(start_node, end_node, keys)] = row[6]

# Open the Excel file
wb = openpyxl.load_workbook(file, data_only=True)

# Get the worksheet with the nodes
ws2 = wb['Nodes']
        
#adding coordinates displaying the graph
pos = {}

for row in ws2.iter_rows(min_row=2, values_only=True):
    pos[row[0]] = (row[2], row[3])

nx.set_node_attributes(g, pos, 'pos')

nx.draw(g, pos=nx.get_node_attributes(g, 'pos'), with_labels=True, connectionstyle='arc3, rad=0.1')
plt.show()

E = list(g.edges())

Ek = []
for u, v, key in g.edges(keys=True):
    Ek.append((u,v,key))
        
n = g.number_of_edges()
N = g.number_of_nodes()
F = sorted(list(g.nodes())) #All nodes in an ordered list

print('Er zijn in totaliteit', n ,'mogelijke trajecten. Deze trajecten verbinden', N, 'knopen.')

#Determining trajects which need to be included as they are essential for the network to be connected
EP = []     #list of the indexes of the essential pipelines in list E

for i in F:
    if g.degree(i) == 1: 
        EP += (E.index((u, v)) for u,v in g.in_edges(i))
        EP +=  (E.index((u, v)) for u,v in g.out_edges(i))
        
if EP == []:
    print('er zijn geen essentiële pijpleidingen in de volledige graaf')
else:
    print('de essentiële pijpleidingen in de volledige graaf zijn:', [E[i] for i in EP])  
    

#%%
"""
=============================================================================
Reading which nodes are sinks and which are sources in the various scenarios
=============================================================================
"""

# Open the Excel file
wb = openpyxl.load_workbook(file, data_only=True)

# Get the worksheet with the nodes
ws2 = wb['Nodes']

def determing_sinkorsource(rownumber):
    demand = {}
    sinks = 0
    sources = 0
    for row in ws2.iter_rows(min_row=2, max_row= max_row_nodes, values_only=True):
        if row[0] in F:
            if row[rownumber] == None:
                demand[row[0]] = 0
            else:
                demand[row[0]] = int(10*row[rownumber])
                if row[rownumber] > 0:
                    sinks += 1
                elif row[rownumber] < 0:
                    sources += 1
    return demand, sinks, sources    

if switch_test == "Yes":
    demand_scenario1, sinks_scenario1, sources_scenario1 = determing_sinkorsource(4)
else:
    for i in demand_scenarios:
        if i == demand_scenarios[0]:
            demand_scenario1, sinks_scenario1, sources_scenario1 = determing_sinkorsource(4)
        elif i == demand_scenarios[1]:
            demand_scenario2, sinks_scenario2, sources_scenario2 = determing_sinkorsource(6)
        elif i == demand_scenarios[2]:
            demand_scenario3, sinks_scenario3, sources_scenario3 = determing_sinkorsource(8)
        elif i == demand_scenarios[3]:
            demand_scenario4, sinks_scenario4, sources_scenario4 = determing_sinkorsource(10)

if switch_test == "Yes":
    print('er zijn',sources_scenario1,'sources en', sinks_scenario1, 'sinks in the variant', demand_scenarios[0])
else:
    print('er zijn',sources_scenario1,'sources en', sinks_scenario1, 'sinks in the variant', demand_scenarios[0])
    print('er zijn',sources_scenario2,'sources en', sinks_scenario2, 'sinks in the variant', demand_scenarios[1])
    print('er zijn',sources_scenario3,'sources en', sinks_scenario3, 'sinks in the variant', demand_scenarios[2])
    print('er zijn',sources_scenario4,'sources en', sinks_scenario4, 'sinks in the variant', demand_scenarios[3])
    

#%%
"""
=============================================================================
Determining variants with whether or not a pipeline exists
=============================================================================
"""

trajects = []

for i in E:
    if E.index(i) in EP:
        trajects.append([1])
    else:
        trajects.append([0,1])

#%%
"""
=============================================================================
Determining variants and calculating costst of and maximal flows through the network
=============================================================================
"""

filename = 'results'
filenumber = '0'

counter = 0

filesave = open((filename+filenumber+'.csv'), 'w', newline='')
writer = csv.writer(filesave, delimiter=',')
writer.writerow(["scenario", "v", "flow_value", "costs", "flowdict"])

x_scenario1 = []
x_scenario2 = []
x_scenario3 = []
x_scenario4 = []

y_scenario1 = []
y_scenario2 = []
y_scenario3 = []
y_scenario4 = []

def determineflows(graph, demand):
    # Make directed graph
    I = graph.to_directed()
    if switch_determining_capacitiesmodel2 == "Yes":
        #Get the worksheet with the edges
        ws3 = wb['Additional_pipelines']
        for row in ws3.iter_rows(min_row=2, max_row= max_row_additionalpipelines, values_only=True):
            I.add_edge(row[0],row[1], length = row[2])
    # Apply maximum flow to find the maximum flow through the network
    I.add_nodes_from(['supersource', 'supersink'])
    for i in F:
        if int(demand[i]) < 0:
            I.add_edge('supersource',i, capacity = -int(demand[i]), length = 1)
        elif int(demand[i]) > 0:
            I.add_edge(i, 'supersink', capacity = int(demand[i]), length = 1)
        flowdict = nx.max_flow_min_cost(I, 'supersource', 'supersink',capacity='capacity', weight = 'length')
    return flowdict

def determine_costs_and_flow(flowdict,v):
    totalflow = 0
    for i in list(flowdict['supersource'].keys()):
        totalflow += int(0.1*flowdict['supersource'][i])
    costs = 0
    length2 = {}
    v1 = []
    [v1.append(0) for i in range(len(Ek))]
    for i in list(flowdict.keys()):
        if i != 'supersource' and i != 'supersink' and i < 100:
            for j in list(flowdict[i].keys()):
                 if j !='supersource' and j != 'supersink' and j < 100:
                     if flowdict[i][j] > 0:
                         if (i,j,0) in Ek:
                             v1[Ek.index((i,j,0))] = 1
                         else:
                             v1[Ek.index((j,i,0))] = 1
                         if (i,j,0) in list(length.keys()):
                             length2[(i,j)] = length[(i,j,0)]
                         else:
                             length2[(i,j)] = length[(j,i,0)]
                         if switch_determining_newpipelines_necessary == "Yes" or switch_determining_capacitiesmodel2:
                             if (i,j,0) in reuse[0]:
                                 costs += length2[(i,j)]/0.4*c2*((0.1*flowdict[i][j])/cap)**Rc
                             elif (j,i,0) in reuse[0]:
                                 costs += length2[(i,j)]/0.4*c2*((0.1*flowdict[i][j])/cap)**Rc
                             else:
                                 costs += length2[(i,j)]*c1*((0.1*flowdict[i][j])/cap)**Rc
                         else:
                             costs += length2[(i,j)]*c1*((0.1*flowdict[i][j])/cap)**Rc
    costs = round((costs/1000000),1)
    return totalflow, costs, v1

for v in itertools.product(*trajects):
    n1 = v.count(1) #number of edges in graph
    if n1 >= N-1:
        G = nx.MultiGraph([Ek[i] for i in range(len(v)) if v[i]== 1]) #Define graph
        G.add_nodes_from(F) #Add all nodes
        l_N = list(G.nodes())
        if nx.is_connected(G):  
            counter += 1
            if counter == 40000:
                print('hello')
                break
            if counter % (1000000/len(demand_scenarios)) == 0:
                filesave.close
                filenumber = str(int(counter/(1000000/len(demand_scenarios))))
                filesave = open((filename+filenumber+'.csv'), 'w', newline='')
                writer = csv.writer(filesave, delimiter=',')                                
            # Length of edges 
            nx.set_edge_attributes(G,length, 'length')
            nx.set_edge_attributes(G,capacities, 'capacity')
            H = nx.Graph()
            for o,p,key,data in G.edges(data=True, keys=True):
                if (o,p) not in H.edges():
                    H.add_edge(o, p, length = data['length'], capacity = data['capacity'])
            if switch_test == "Yes":
                scenario = demand_scenarios[0]
                flowdict_scenario1 = determineflows(H, demand_scenario1)
                totalflow, costs, v1 = determine_costs_and_flow(flowdict_scenario1,v)
                y_scenario1.append(totalflow)
                x_scenario1.append(costs)
                writer.writerow([scenario,v1,totalflow, costs, flowdict_scenario1])
            else:
                for scenario in demand_scenarios:
                     if scenario == demand_scenarios[0]:
                         flowdict_scenario1 = determineflows(H, demand_scenario1)
                         totalflow, costs, v1 = determine_costs_and_flow(flowdict_scenario1,v)
                         y_scenario1.append(totalflow)
                         x_scenario1.append(costs)
                         writer.writerow([scenario,v1,totalflow, costs, flowdict_scenario1])
                     elif scenario == demand_scenarios[1]:
                         flowdict_scenario2 = determineflows(H, demand_scenario2)
                         totalflow, costs, v1 = determine_costs_and_flow(flowdict_scenario2,v)
                         y_scenario2.append(totalflow)
                         x_scenario2.append(costs)
                         writer.writerow([scenario,v1,totalflow, costs, flowdict_scenario2])
                     elif scenario == demand_scenarios[2]:
                         flowdict_scenario3 = determineflows(H, demand_scenario3)
                         totalflow, costs, v1 = determine_costs_and_flow(flowdict_scenario3,v)
                         y_scenario3.append(totalflow)
                         x_scenario3.append(costs)
                         writer.writerow([scenario,v1,totalflow, costs,flowdict_scenario3])
                     elif scenario == demand_scenarios[3]:
                         flowdict_scenario4 = determineflows(H, demand_scenario4)
                         totalflow, costs, v1 = determine_costs_and_flow(flowdict_scenario4,v)
                         y_scenario4.append(totalflow)
                         x_scenario4.append(costs)
                         writer.writerow([scenario,v1,totalflow, costs,flowdict_scenario4])

filesave.close()

#%%
"""
=============================================================================
Determining the highest flows over each of the routes and compare this 
flow to the reuse capacity to determine whether or not reuse is sufficient
to meet the demand
=============================================================================
"""

if switch_determining_newpipelines_necessary == 'Yes':
    
    necessary_capacities = {}
    nodelist = []
    
    for row in ws1.iter_rows(min_row=2, values_only=True):
        if (row[0],row[1]) not in list(necessary_capacities.keys()):
            necessary_capacities[(row[0],row[1])] = 0
    
    with open((filename+filenumber+'.csv'), 'r', newline='') as results:
        reader = csv.reader(results)
        next(reader, None)
        for row in reader:
            dictionary_flowvalues = ast.literal_eval(row[4])
            for startnode in F:
                for endnode in F:
                    if (startnode,endnode) in list(necessary_capacities.keys()):
                        if startnode in list(dictionary_flowvalues.keys()):
                            if endnode in list(dictionary_flowvalues[startnode].keys()):
                                if necessary_capacities[(startnode, endnode)] < round(0.1*dictionary_flowvalues[startnode][endnode],1):
                                    necessary_capacities[(startnode, endnode)] = round(0.1*dictionary_flowvalues[startnode][endnode],1)
                        if endnode in list(dictionary_flowvalues.keys()): 
                            if startnode in list(dictionary_flowvalues[endnode].keys()):
                                if necessary_capacities[(startnode, endnode)] < round(0.1*dictionary_flowvalues[endnode][startnode],1):
                                    necessary_capacities[(startnode, endnode)] = round(0.1*dictionary_flowvalues[endnode][startnode],1)
else:
    pass
                            
if switch_determining_newpipelines_necessary == 'Yes':   
    
    filename_necessarypipelines = 'necessarypipelines'
    
    filesave = open((filename_necessarypipelines+modelyear+'.csv'), 'w', newline='')
    writer = csv.writer(filesave, delimiter=',')
    writer.writerow(["route", "necessary capacity", "reuse capacity", "new pipeline necessary?"])
    for edge in list(necessary_capacities.keys()):
        if (edge[0],edge[1],0) in reuse[0]:
            reuse_index = reuse[0].index((edge[0],edge[1],0))
            reuse_capacity = reuse[2][reuse_index]
            if reuse[2][reuse_index] < necessary_capacities[edge]:
                new_necessary = True
            else:
                new_necessary = False
            writer.writerow([edge,necessary_capacities[edge],reuse_capacity,new_necessary]) 
        elif (edge[1],edge[0],0) in reuse[0]:
            reuse_index = reuse[0].index((edge[1],edge[0],0))
            reuse_capacity = reuse[2][reuse_index]
            if reuse[2][reuse_index] < necessary_capacities[edge]:
                new_necessary = True
            else:
                new_necessary = False
            writer.writerow([edge,necessary_capacities[edge],reuse_capacity,new_necessary]) 
        else:
            writer.writerow([edge,necessary_capacities[edge],'NA','NA'])
else:
    pass

filesave.close()

#%%
"""
=============================================================================
Creating an overview of the flows over each of the pipelines. This 
information can be used to manually determine the capacities that are assigned 
to the pipelines in the following steps of the model
=============================================================================
"""

if switch_determining_capacitiesmodel2 == 'Yes':
    
    index_edges = []

    for row in ws1.iter_rows(min_row=2, max_row = max_row_pipelines, values_only=True):
        if (row[0],row[1]) not in index_edges:
            index_edges.append((row[0],row[1]))
    
    ws3 = wb['Additional_pipelines']
    for row in ws3.iter_rows(min_row=2, max_row= max_row_additionalpipelines, values_only=True):
        if (row[0],row[1]) not in index_edges:
            index_edges.append((row[0],row[1]))
            
    filename = 'results'
    filenumber = '0'
    
    capacities_model2_scenario1 = []
    capacities_model2_scenario2 = []
    capacities_model2_scenario3 = []
    capacities_model2_scenario4 = []
    
    
    for i in range(len(index_edges)):
        capacities_model2_scenario1.append([])
        capacities_model2_scenario2.append([])
        capacities_model2_scenario3.append([])
        capacities_model2_scenario4.append([])
        
    def determine_capacitiesmodel2(capacities_model2):
        dictionary_flowvalues = ast.literal_eval(row[4])
        for startnode in list(dictionary_flowvalues.keys()):
            for endnode in list(dictionary_flowvalues[startnode].keys()):
                value = dictionary_flowvalues[startnode][endnode]/10
                if (startnode, endnode) in index_edges:
                    number = index_edges.index((startnode, endnode))
                    if value not in capacities_model2[number]:
                        capacities_model2[number].append(value)
                elif (endnode, startnode) in index_edges:
                    number = index_edges.index((endnode, startnode))
                    if value not in capacities_model2[number]:
                        capacities_model2[number].append(value)
        return capacities_model2
    
    with open((filename+filenumber+'.csv'), 'r', newline='') as results:
        reader = csv.reader(results)
        next(reader, None)
        for row in reader:
            if row[0] == demand_scenarios[0]:
                capacities_model2_scenario1 = determine_capacitiesmodel2(capacities_model2_scenario1)
            elif row[0] == demand_scenarios[1]:
                capacities_model2_scenario2 = determine_capacitiesmodel2(capacities_model2_scenario2)
            elif row[0] == demand_scenarios[2]:
                capacities_model2_scenario3 = determine_capacitiesmodel2(capacities_model2_scenario3)
            elif row[0] == demand_scenarios[3]:
                capacities_model2_scenario4 = determine_capacitiesmodel2(capacities_model2_scenario4)
    
    filesave = open(('capaciteitenmodel2_'+modelyear+'.csv'), 'w', newline='')
    writer = csv.writer(filesave, delimiter=',')
    writer.writerow(["Van", "Naar", "mogelijke capaciteiten"])
        
    for scenario in demand_scenarios:
        for i in range(len(index_edges)):
            if scenario == demand_scenarios[0]:
                capacities_model2_scenario1[i] = sorted(capacities_model2_scenario1[i])
            elif scenario == demand_scenarios[1]:
                capacities_model2_scenario2[i] = sorted(capacities_model2_scenario2[i])
            elif scenario == demand_scenarios[2]:
                capacities_model2_scenario3[i] = sorted(capacities_model2_scenario3[i])
            elif scenario == demand_scenarios[3]:
                capacities_model2_scenario4[i] = sorted(capacities_model2_scenario4[i])
    
        for i in range(len(index_edges)):
            if scenario == demand_scenarios[0]:
                writer.writerow([scenario, index_edges[i][0],index_edges[i][1],capacities_model2_scenario1[i]])
            elif scenario == demand_scenarios[1]:
                writer.writerow([scenario, index_edges[i][0],index_edges[i][1],capacities_model2_scenario2[i]])
            elif scenario == demand_scenarios[2]:
                writer.writerow([scenario, index_edges[i][0],index_edges[i][1],capacities_model2_scenario3[i]])
            elif scenario == demand_scenarios[3]:
                writer.writerow([scenario, index_edges[i][0],index_edges[i][1],capacities_model2_scenario4[i]])
    
    filesave.close()
    

#%%
"""
=============================================================================
Timing the calculations and make a sound
=============================================================================
"""

endtime = time.time()
print('de berekeningen duurden', round((endtime-starttime)), 'seconden.')

#make a sound when done
print('\a')